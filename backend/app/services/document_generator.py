"""
ドキュメント生成サービス
Excel形式のInvoiceとPacking Listを生成
"""
from pathlib import Path
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.case import Case
from app.models.customer import Customer
from app.models.product import Product
from app.models.document import Document
from app.models.user import User


class DocumentGenerator:
    """ドキュメント生成クラス"""

    def __init__(self, db: Session):
        self.db = db
        self.templates_dir = Path(__file__).parent.parent.parent / "templates"
        self.output_dir = Path(__file__).parent.parent.parent / "generated_documents"

    def generate_invoice(
        self,
        case_id: int,
        user_id: int,
        template_name: Optional[str] = None
    ) -> Document:
        """
        Invoice（請求書）を生成

        Args:
            case_id: 案件ID
            user_id: 生成者ID
            template_name: テンプレート名（オプション）

        Returns:
            Document: 生成されたドキュメントのレコード
        """
        # 案件情報を取得
        case = self.db.query(Case).filter(Case.id == case_id).first()
        if not case:
            raise ValueError(f"Case ID {case_id} not found")

        # 顧客情報を取得
        customer = self.db.query(Customer).filter(Customer.id == case.customer_id).first()

        # 商品情報を取得
        product = self.db.query(Product).filter(Product.id == case.product_id).first()

        # Excelワークブック作成
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # タイトル
        ws['A1'] = 'INVOICE'
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')

        # 会社情報（仮）
        ws['A3'] = '発行者：ペンギンロジスティクス株式会社'
        ws['A4'] = '〒100-0001 東京都千代田区千代田1-1-1'
        ws['A5'] = 'TEL: 03-1234-5678 / FAX: 03-1234-5679'

        # 顧客情報
        ws['E3'] = '請求先:'
        ws['E3'].font = Font(bold=True)
        ws['E4'] = customer.customer_name if customer else ""
        ws['E5'] = customer.address if customer and customer.address else ""
        ws['E6'] = customer.contact_person if customer and customer.contact_person else ""

        # 請求書情報
        ws['A8'] = 'Invoice No:'
        ws['B8'] = case.case_number
        ws['B8'].font = Font(bold=True)

        ws['A9'] = 'Invoice Date:'
        ws['B9'] = datetime.now().strftime('%Y-%m-%d')

        ws['A10'] = 'Shipment Date:'
        ws['B10'] = case.shipment_date.strftime('%Y-%m-%d') if case.shipment_date else ""

        # 明細ヘッダー
        header_row = 12
        headers = ['No.', 'Product Code', 'Product Name', 'Quantity', 'Unit', 'Unit Price', 'Amount']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center')

        # 明細データ
        detail_row = header_row + 1
        ws.cell(row=detail_row, column=1).value = 1
        ws.cell(row=detail_row, column=2).value = product.product_code if product else ""
        ws.cell(row=detail_row, column=3).value = product.product_name if product else ""
        ws.cell(row=detail_row, column=4).value = case.quantity
        ws.cell(row=detail_row, column=5).value = case.unit
        ws.cell(row=detail_row, column=6).value = case.sales_unit_price
        ws.cell(row=detail_row, column=7).value = case.sales_amount

        # 明細セルのスタイル設定
        for col_idx in range(1, 8):
            cell = ws.cell(row=detail_row, column=col_idx)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col_idx >= 4:  # 数値列は右寄せ
                cell.alignment = Alignment(horizontal='right')

        # 合計
        total_row = detail_row + 2
        ws.cell(row=total_row, column=6).value = 'Total Amount:'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=total_row, column=7).value = case.sales_amount
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=7).number_format = '#,##0.00'

        # 備考
        if case.notes:
            notes_row = total_row + 2
            ws.cell(row=notes_row, column=1).value = 'Notes:'
            ws.cell(row=notes_row, column=1).font = Font(bold=True)
            ws.cell(row=notes_row + 1, column=1).value = case.notes
            ws.merge_cells(f'A{notes_row + 1}:G{notes_row + 1}')

        # 列幅調整
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        # ファイル保存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"invoice_{case.case_number}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        # 出力ディレクトリが存在しない場合は作成
        self.output_dir.mkdir(parents=True, exist_ok=True)

        wb.save(str(filepath))

        # データベースに記録
        document = Document(
            case_id=case_id,
            document_type="invoice",
            file_name=filename,
            file_path=str(filepath),
            template_name=template_name or "default_invoice",
            generated_by=user_id,
            notes=f"Invoice generated for case {case.case_number}"
        )
        self.db.add(document)
        self.db.commit()
        self.db.refresh(document)

        return document

    def generate_packing_list(
        self,
        case_id: int,
        user_id: int,
        template_name: Optional[str] = None
    ) -> Document:
        """
        Packing List（梱包リスト）を生成

        Args:
            case_id: 案件ID
            user_id: 生成者ID
            template_name: テンプレート名（オプション）

        Returns:
            Document: 生成されたドキュメントのレコード
        """
        # 案件情報を取得
        case = self.db.query(Case).filter(Case.id == case_id).first()
        if not case:
            raise ValueError(f"Case ID {case_id} not found")

        # 顧客情報を取得
        customer = self.db.query(Customer).filter(Customer.id == case.customer_id).first()

        # 商品情報を取得
        product = self.db.query(Product).filter(Product.id == case.product_id).first()

        # Excelワークブック作成
        wb = Workbook()
        ws = wb.active
        ws.title = "Packing List"

        # タイトル
        ws['A1'] = 'PACKING LIST'
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')

        # 会社情報
        ws['A3'] = '発行者：ペンギンロジスティクス株式会社'
        ws['A4'] = '〒100-0001 東京都千代田区千代田1-1-1'
        ws['A5'] = 'TEL: 03-1234-5678 / FAX: 03-1234-5679'

        # 顧客情報
        ws['E3'] = '送付先:'
        ws['E3'].font = Font(bold=True)
        ws['E4'] = customer.customer_name if customer else ""
        ws['E5'] = customer.address if customer and customer.address else ""

        # Packing List情報
        ws['A8'] = 'Packing List No:'
        ws['B8'] = case.case_number
        ws['B8'].font = Font(bold=True)

        ws['A9'] = 'Packing Date:'
        ws['B9'] = datetime.now().strftime('%Y-%m-%d')

        ws['A10'] = 'Shipment Date:'
        ws['B10'] = case.shipment_date.strftime('%Y-%m-%d') if case.shipment_date else ""

        # 明細ヘッダー
        header_row = 12
        headers = ['No.', 'Product Code', 'Product Name', 'Quantity', 'Unit', 'Gross Weight (kg)', 'Net Weight (kg)', 'Packages']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center')

        # 明細データ
        detail_row = header_row + 1
        ws.cell(row=detail_row, column=1).value = 1
        ws.cell(row=detail_row, column=2).value = product.product_code if product else ""
        ws.cell(row=detail_row, column=3).value = product.product_name if product else ""
        ws.cell(row=detail_row, column=4).value = case.quantity
        ws.cell(row=detail_row, column=5).value = case.unit
        # 重量・個数は仮データ（実際の案件には含まれていない場合があるため）
        ws.cell(row=detail_row, column=6).value = case.quantity * 10 if case.quantity else 0  # 仮の総重量
        ws.cell(row=detail_row, column=7).value = case.quantity * 9 if case.quantity else 0  # 仮の正味重量
        ws.cell(row=detail_row, column=8).value = int(case.quantity / 100) if case.quantity else 1  # 仮の個数

        # 明細セルのスタイル設定
        for col_idx in range(1, 9):
            cell = ws.cell(row=detail_row, column=col_idx)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col_idx >= 4:  # 数値列は右寄せ
                cell.alignment = Alignment(horizontal='right')

        # 合計
        total_row = detail_row + 2
        ws.cell(row=total_row, column=3).value = 'Total:'
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=4).value = case.quantity
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        ws.cell(row=total_row, column=6).value = case.quantity * 10 if case.quantity else 0
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7).value = case.quantity * 9 if case.quantity else 0
        ws.cell(row=total_row, column=7).font = Font(bold=True)

        # 備考
        if case.notes:
            notes_row = total_row + 2
            ws.cell(row=notes_row, column=1).value = 'Notes:'
            ws.cell(row=notes_row, column=1).font = Font(bold=True)
            ws.cell(row=notes_row + 1, column=1).value = case.notes
            ws.merge_cells(f'A{notes_row + 1}:H{notes_row + 1}')

        # 列幅調整
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12

        # ファイル保存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"packing_list_{case.case_number}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        # 出力ディレクトリが存在しない場合は作成
        self.output_dir.mkdir(parents=True, exist_ok=True)

        wb.save(str(filepath))

        # データベースに記録
        document = Document(
            case_id=case_id,
            document_type="packing_list",
            file_name=filename,
            file_path=str(filepath),
            template_name=template_name or "default_packing_list",
            generated_by=user_id,
            notes=f"Packing list generated for case {case.case_number}"
        )
        self.db.add(document)
        self.db.commit()
        self.db.refresh(document)

        return document

    def get_documents(
        self,
        case_id: Optional[int] = None,
        document_type: Optional[str] = None,
        skip: int = 0,
        limit: int = 100
    ) -> tuple[list[Document], int]:
        """
        ドキュメント一覧を取得

        Args:
            case_id: 案件IDでフィルタリング（オプション）
            document_type: ドキュメントタイプでフィルタリング（オプション）
            skip: スキップ数
            limit: 取得件数上限

        Returns:
            tuple: (ドキュメントリスト, 総件数)
        """
        query = self.db.query(Document)

        if case_id:
            query = query.filter(Document.case_id == case_id)

        if document_type:
            query = query.filter(Document.document_type == document_type)

        total = query.count()
        documents = query.order_by(Document.generated_at.desc()).offset(skip).limit(limit).all()

        return documents, total

    def get_document_file(self, document_id: int) -> Path:
        """
        ドキュメントファイルのパスを取得

        Args:
            document_id: ドキュメントID

        Returns:
            Path: ファイルパス
        """
        document = self.db.query(Document).filter(Document.id == document_id).first()
        if not document:
            raise ValueError(f"Document ID {document_id} not found")

        if not document.file_path:
            raise ValueError(f"Document ID {document_id} has no file path")

        filepath = Path(document.file_path)
        if not filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")

        return filepath
